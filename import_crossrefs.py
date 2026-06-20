"""
import_crossrefs.py
──────────────────
Reads MatchIT_Key_Ref_File_.xlsx (or any similarly structured file),
matches every SKU in the images database against all brand columns,
and writes sku_crossrefs.json.

SKUs not yet in the DB are stored anyway — they'll activate automatically
when images are added later.

Run from your project folder:
    python import_crossrefs.py

Options:
    python import_crossrefs.py --xlsx "C:\\path\\to\\file.xlsx"
    python import_crossrefs.py --db "C:\\path\\to\\images.db"
    python import_crossrefs.py --out my_crossrefs.json
    python import_crossrefs.py --all   (import ALL rows, not just DB matches)
"""

import argparse
import json
import sqlite3
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: Run:  pip install openpyxl")
    sys.exit(1)

# ── Defaults ─────────────────────────────────────────────────────────────────
DEFAULT_XLSX = "MatchIT_Key_Ref_File_.xlsx"
DEFAULT_JSON = "sku_crossrefs.json"
NULL_VALUES  = {"NULL", "NONE", "", "-", "N/A"}
# ─────────────────────────────────────────────────────────────────────────────


def load_excel(xlsx_path: str):
    """
    Returns:
        headers      : list of column headers
        brand_headers: list of (col_index, brand_name) for cols 2 onwards
        lookup       : dict of code.upper() -> row_number
        ws           : the worksheet object
    """
    print(f"Loading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    print(f"  {ws.max_row-1} rows, {ws.max_column} columns")

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    print(f"  Columns: {headers}")

    # Brand columns = col 3 onwards (col 1=Manufacturer, col 2=Name/primary ref — excluded from crossrefs)
    brand_headers = [(col, headers[col-1]) for col in range(3, ws.max_column + 1)]

    # Build reverse lookup using ALL columns including Name (col 2) for matching
    lookup_cols = [(col, headers[col-1]) for col in range(2, ws.max_column + 1)]
    lookup = {}
    for row in range(2, ws.max_row + 1):
        for col, _ in lookup_cols:
            val = ws.cell(row, col).value
            if val:
                clean = str(val).strip()
                if clean.upper() not in NULL_VALUES:
                    key = clean.upper()
                    if key not in lookup:
                        lookup[key] = row

    print(f"  Built lookup: {len(lookup)} unique reference codes")
    return headers, brand_headers, lookup, ws


def build_crossrefs_for_row(ws, row, brand_headers):
    """Extract all non-null brand references from a row."""
    refs = []
    for col, brand_name in brand_headers:
        val = ws.cell(row, col).value
        if val:
            clean = str(val).strip()
            if clean.upper() not in NULL_VALUES:
                refs.append({"brand": brand_name, "code": clean})
    return refs


def get_db_skus(db_path: str):
    """Return list of distinct SKUs from the images database."""
    try:
        conn = sqlite3.connect(db_path)
        skus = [r[0] for r in conn.execute("SELECT DISTINCT sku FROM images").fetchall()]
        conn.close()
        return skus
    except Exception as e:
        print(f"  [WARN] Could not read DB: {e}")
        return []


def main():
    parser = argparse.ArgumentParser(description="Import cross-references from Excel into sku_crossrefs.json")
    parser.add_argument("--xlsx", default=DEFAULT_XLSX,
                        help=f"Path to Excel file (default: {DEFAULT_XLSX})")
    parser.add_argument("--db",   default=None,
                        help="Path to images.db (auto-detected if not specified)")
    parser.add_argument("--out",  default=DEFAULT_JSON,
                        help=f"Output JSON file (default: {DEFAULT_JSON})")
    parser.add_argument("--all",  action="store_true",
                        help="Import ALL rows from Excel, not just ones matching DB SKUs")
    args = parser.parse_args()

    # Find Excel file
    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        # Try searching current folder
        found = list(Path(".").glob("*.xlsx"))
        if found:
            xlsx_path = found[0]
            print(f"Found Excel: {xlsx_path}")
        else:
            print(f"ERROR: Cannot find {args.xlsx}")
            sys.exit(1)

    # Find DB
    db_path = args.db
    if not db_path:
        candidates = [
            r"C:\Users\c_a_b\AppData\Local\MatchITv2_ProductMatch_Data\images.db",
            "images.db", "keys.db"
        ]
        for c in candidates:
            if Path(c).exists():
                db_path = c
                break

    # Load Excel
    headers, brand_headers, lookup, ws = load_excel(str(xlsx_path))

    result = {}

    if args.all:
        # Import every row regardless of DB
        print("\nImporting ALL rows from Excel...")
        for row in range(2, ws.max_row + 1):
            # Use the Name column (col 2) as the SKU key
            name_val = ws.cell(row, 2).value
            if not name_val:
                continue
            sku_key = str(name_val).strip()
            if sku_key.upper() in NULL_VALUES:
                continue
            mfr  = ws.cell(row, 1).value or ""
            refs = build_crossrefs_for_row(ws, row, brand_headers)
            if refs:
                result[sku_key] = {
                    "manufacturer": str(mfr).strip(),
                    "crossrefs": refs
                }
        print(f"  Imported {len(result)} entries")

    else:
        # Match against DB SKUs first, then import unmatched Excel rows too
        db_skus = []
        if db_path and Path(db_path).exists():
            db_skus = get_db_skus(db_path)
            print(f"\nDB SKUs found: {len(db_skus)}")
        else:
            print("\n[WARN] No DB found — importing all Excel rows by Name column")
            args.all = True

        matched   = []
        unmatched = []

        for sku in db_skus:
            row = lookup.get(sku.upper())
            if row:
                mfr  = ws.cell(row, 1).value or ""
                refs = build_crossrefs_for_row(ws, row, brand_headers)
                result[sku] = {
                    "manufacturer": str(mfr).strip(),
                    "crossrefs": refs
                }
                matched.append(sku)
            else:
                unmatched.append(sku)

        print(f"  Matched:   {len(matched)} SKUs")
        if unmatched:
            print(f"  Unmatched: {len(unmatched)} SKUs (no Excel entry found)")
            print(f"             {unmatched}")

        # Also store ALL Excel rows so future DB additions auto-resolve
        print("\nStoring remaining Excel rows for future SKUs...")
        future_count = 0
        for row in range(2, ws.max_row + 1):
            name_val = ws.cell(row, 2).value
            if not name_val:
                continue
            name_clean = str(name_val).strip()
            if name_clean.upper() in NULL_VALUES:
                continue
            # Skip if already added via DB match
            if name_clean in result:
                continue
            # Check if any code in this row matches an already-added SKU
            already_added = False
            for col, _ in brand_headers:
                v = ws.cell(row, col).value
                if v and str(v).strip().upper() in [s.upper() for s in result.keys()]:
                    already_added = True
                    break
            if already_added:
                continue

            mfr  = ws.cell(row, 1).value or ""
            refs = build_crossrefs_for_row(ws, row, brand_headers)
            if refs:
                result[name_clean] = {
                    "manufacturer": str(mfr).strip(),
                    "crossrefs": refs
                }
                future_count += 1

        print(f"  Stored {future_count} additional rows for future SKUs")

    # Write JSON
    out_path = Path(args.out)
    out_path.write_text(
        json.dumps(result, indent=2, ensure_ascii=False),
        encoding="utf-8"
    )

    total_refs = sum(len(v["crossrefs"]) for v in result.values())
    print(f"\n{'-'*55}")
    print(f"Done — {len(result)} SKU entries, {total_refs} total cross-references")
    print(f"Output: {out_path.resolve()}")
    print(f"\nDeploy {out_path} to your project root and restart Flask.")


if __name__ == "__main__":
    main()